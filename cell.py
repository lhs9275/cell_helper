import pandas as pd
import re
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

# 엑셀 파일
print("안녕")

change_colums = 0
df_data = pd.read_excel('예약.xlsx', index_col=0)
new_column_names = ['1코트', '2코트', '3코트','4코트','5코트','6코트','7코트','8코트']  # 필요한 만큼 열 이름을 변경
new_index_values = ['06:00~07:00', '07:00~08:00', '08:00~09:00', '09:00~10:00', '10:00~11:00', '11:00~12:00', '12:00~13:00','13:00~14:00','14:00~15:00','15:00~16:00','16:00~17:00','17:00~18:00','18:00~19:00','19:00~20:00','20:00~21:00','21:00~22:00'] #행


#조건 리스트
desired_reservation_status_list = ['결제가능', '상담대기', '예약완료']
desired_facility_list= ['안성맞춤소프트테니스구장(테니스구장(1코트))','안성맞춤소프트테니스구장(테니스구장(2코트))','안성맞춤소프트테니스구장(테니스구장(3코트))','안성맞춤소프트테니스구장(테니스구장(4코트))','안성맞춤소프트테니스구장(테니스구장(5코트))','안성맞춤소프트테니스구장(테니스구장(6코트))','안성맞춤소프트테니스구장(테니스구장(7코트))','안성맞춤소프트테니스구장(테니스구장(8코트))']
desired_reservation_time_list = ['06:00~08:00', '08:00~10:00', '10:00~12:00', '12:00~14:00', '14:00~16:00', '16:00~18:00', '18:00~20:00', '20:00~22:00']
desired_reservation_time_list_4 = ['06:00~10:00', '08:00~12:00', '10:00~14:00', '12:00~16:00', '14:00~18:00', '16:00~20:00', '18:00~22:00']
desired_reservation_time_list_r = ['6-8','8-10','10-12','12-14','14-16','16-18','18-20','20-22']
desired_reservation_time_list_4_r = ['6-10', '8-12', '10-14', '12-16', '14-18', '16-20', '18-22']

# 새로운 엑셀 파일을 생성
df_sch = pd.DataFrame(index=new_index_values, columns=new_column_names)

#조건문 

for change_colums in range(8):
    j=0
    j_1=0
    for i in range (8):
        if (df_data['시설명'] == desired_facility_list[change_colums]).any():
            condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))

            if condition.any():
                # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                reserved_member = condition[condition].index[0]
                combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]}"
                df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value
        j=j+2

    for k in range(7):
        if (df_data['시설명'] == desired_facility_list[change_colums]).any():
            condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))
            if condition.any():
                # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                reserved_member = condition[condition].index[0]
                combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]}"
                df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value
        j_1=j_1+2





def remove_parentheses(value):
    return re.sub(r'\([^)]*\)', '', str(value))


df_sch = df_sch.applymap(remove_parentheses)

df_sch = df_sch.replace('nan', '')

# ExcelWriter 객체 생성
with pd.ExcelWriter('tennis_court_schedule.xlsx', engine='openpyxl') as writer:
    # DataFrame을 Excel 파일에 쓰기
    df_sch.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=1, header=True, index=True)

    # ExcelWriter 객체에서 워크북과 워크시트 객체 가져오기
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # 열의 너비를 15로 설정
    for col_num, value in enumerate(new_column_names):
        worksheet.column_dimensions[worksheet.cell(row=2, column=col_num+1).column_letter].width = 13

    # 높이 설정
    for row_num, value in enumerate(new_index_values):
        worksheet.row_dimensions[row_num + 3].height = 25

    # 헤더 텍스트 추가
    header_text = "                            정구장 (클레이 코트)             월     일    요일"  # 페이지 번호를 나타내는 예시
    worksheet['A1'] = header_text
    worksheet.merge_cells('A1:I1')

    # A1 셀의 높이를 늘리기
    worksheet.row_dimensions[1].height = 30

    worksheet['A1'].font = Font(size=16)
    # 외곽에 선 추가
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    # 안쪽에 선 추가
    inside_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin', color='000000'))

    # 셀에 스타일 적용
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(size=9)

    # 안쪽에 선 추가
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = inside_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(size=9)

    worksheet['A1'].font = Font(size=16, bold=True)

print("엑셀 파일이 생성되었습니다.")
